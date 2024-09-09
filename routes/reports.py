import pandas as pd
from fastapi import APIRouter
from fastapi.responses import JSONResponse, FileResponse
import psycopg2
from pathlib import Path
from typing import Optional
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import os
import json

from fastapi.responses import JSONResponse, FileResponse
from fastapi import FastAPI, Query, HTTPException
from datetime import date, datetime
from config.database import cursor
from fpdf import FPDF
import textwrap

from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib import colors
from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader

app = FastAPI()
reports = APIRouter()

DirectoryEmpleados = "reportsfile/administracion/empleados/"

pdfmetrics.registerFont(TTFont('Roboto', 'fonts/Roboto/Roboto-Regular.ttf'))
pdfmetrics.registerFont(TTFont('Roboto-Bold', 'fonts/Roboto/Roboto-Bold.ttf'))


# Validar si la carpeta ya existe
if not os.path.exists(DirectoryEmpleados):
    # Si no existe, crear la carpeta
    os.makedirs(DirectoryEmpleados)

now = date.today()

@reports.get('/empleados', tags=["ReportsXls"])
def getEmpleados():
    return {"message": "empleados"}

# Users #       #
tituloModulo = "Reporte de Usuarios"
def getUsuarios():
    query = """
	        select  
            e.name as "Nombre", 
            e.email as "Correo Electrónico", 
            string_agg(distinct r.title, ', ' ) as "Roles",
            e.name as "Empleado Vinculado",
            a.area as "Área", 
            string_agg(distinct p.puesto, ', ' ) as "Puesto"
            from empleados e  
            inner join role_user ru on e.id=ru.user_id 
            inner join roles r on ru.role_id =r.id
            inner join areas a ON e.area_id=a.id
            inner join puestos p ON e.puesto_id =p.id
            where r.deleted_at is null
            group  by e.name, e.email, a.area
            order by e.name asc
        """

    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@reports.get('/usuarios/pdf', tags=["ReportsPDF"])
def getUsuariosPDF():
    resultados = getUsuarios()

    pdf_filename = f"usuarios_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["Nombre", "Correo \nElectrónico", "Roles", "Empleado  \nVinculado", "Área", "Puesto"]
    tam_celdas =[1.5 * inch ]

    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')


@reports.get('/usuarios', tags=["ReportsXls"])
def getUsuariosExcel():
    resultados = getUsuarios()
    fileRoute = DirectoryEmpleados + "usuarios_" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(status_code=404, detail="file not found on the server")

    return FileResponse(excel_path, filename=f"usuarios_{now}.xlsx", media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# Empleados Puestos  #
tituloModulo = "Reporte de Ompleados-Puestos"
def getEmpleadosPuestos():
    query = """
            select 
            e.name as "Empleado",
            s.name as "Supervisor",
            a.area as "Área",
            p.puesto as "Puesto"
            from empleados e
            left join empleados s on e.supervisor_id=s.id
            inner join areas a on e.area_id=a.id
            inner join  puestos p on e.puesto_id =p.id
            where e.deleted_at is null and e.estatus='alta'
            order by e.name asc
        """
    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados


@reports.get('/empleadosPuestos', tags=["ReportsXls"])
def getEmpleadosPuestosExcel():
    resultados=getEmpleadosPuestos()
    fileRoute = DirectoryEmpleados + "empleadosPuestos" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

@reports.get('/empleadosPuestos/pdf', tags=["ReportsPDF"])
def getEmpleadosPdf():
    resultados = getEmpleadosPuestos()

    pdf_filename = f"empleadosPuestos_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["Empleado", "Supervisor", "Área", "Puesto"]
    tam_celdas =[1.5 * inch ]

    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')

    
## Puestos  #
tituloModulo = "Reporte Módulo Puestos"
def getPuestos():
    query = """
            select 
            p.puesto as "Puesto", 
            a.area as "Área", 
            a.descripcion as "Descripción" 
            from puestos p 
            inner join 
            areas a on p.id_area=a.id            
            where p.deleted_at is null 
        """
    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@reports.get('/moduloPuestos', tags=["ReportsXls"])
def getEmpleadosPuestosExcel():
    resultados=getPuestos
    fileRoute = DirectoryEmpleados + "puestos" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(
            status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

@reports.get('/moduloPuestos/pdf', tags=["ReportsPDF"])
def getEmpleadosPuestosPDF():   
    resultados = getPuestos()

    pdf_filename = f"puestos_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["Puesto", "Área", "Descripción"]
    tam_celdas =[1.5 * inch ]

    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')

    
## Roles  #
tituloModulo = "Reporte Módulo Roles"
def getRoles():
    query = """
            select 
                r.id as "ID", 
                r.title as "Nombre del rol"
            from roles r 
            where r.deleted_at is null;
        """
    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@reports.get('/moduloRoles', tags=["ReportsXls"])
def getModuloRolesExcel():
    resultados = getRoles()
    fileRoute = DirectoryEmpleados + "moduloRoles-" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(
            status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)


@reports.get('/moduloRoles/pdf', tags=["ReportsPDF"])
def getModuloRolesPDF():
    resultados = getRoles()

    pdf_filename = f"moduloRoles_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["ID", "CNombre \ndel Rol"]
    tam_celdas =[1.5 * inch ]

    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')


## Soporte #
tituloModulo = "Reporte Módulo Soporte"
def getSoporte():
    query = """
            select 
            cs.id as "ID",
            cs.rol as "Rol",
            e.name as "Nombre",
            p.puesto as "Puesto", 
            cs.telefono as "Teléfono", 
            cs.extension as "Extensión",
            cs.tel_celular as "Tel. Celular",
            cs.correo as "Correo"
            from configuracion_soporte cs 
            inner join empleados e on cs.id_elaboro=e.id 
            inner join puestos p on e.puesto_id =p.id 
            where cs.deleted_at is null
        """
    resultados = ejecutar_consulta_sql(cursor, query)

@reports.get('/soporte', tags=["ReportsXls"])
def getSoporteExcel():
    resultados = getSoporte
    fileRoute = DirectoryEmpleados + "soporte-" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")
    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

@reports.get('/soporte/pdf', tags=["ReportsPDF"])
def getSoportePDF():
    resultados = getSoporte()
    pdf_filename = f"soporte_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["ID", "Rol", "Nombre", "Puesto", "Teléfono", "Extensión", "Tel.Celular", "Correo"]
    tam_celdas =[1.5 * inch ]

    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')

    
## Modulo Empleados  #
tituloModulo = "Reporte Módulo Empleados"
def getModuloEmpleados():
    query = """
            select 
            e.n_empleado as "No.Empleado",
            e.name as "Nombre",
            e.email as "Email",
            e.telefono as "Teléfono",
            a.area as "Área",
            p.puesto as "Puesto",
            s.name as "Supervisor",
            e.antiguedad as "Antigüedad",
            e.estatus as "Estatus"
            from empleados e
            left join empleados s on e.supervisor_id=s.id
            inner join areas a on e.area_id=a.id
            inner join  puestos p on e.puesto_id =p.id
            where e.deleted_at is null and e.estatus='alta'
            order by e.name asc 
        """
    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@reports.get('/moduloEmpleados', tags=["ReportsXls"])
def getModEmpleadosExcel():
    resultados = getModuloEmpleados()
    fileRoute = DirectoryEmpleados + "empleados-" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

@reports.get('/moduloEmpleados/pdf', tags=["ReportsPDF"])
def getModEmpleadosPDF():
    resultados = getModuloEmpleados()

    pdf_filename = f"moduloEmpleados_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["No.Empleado", "Nombre", "Email", "Teléfono", "Área", "Puesto", "Supervisor", "Antigüedad", "Estatus"]
    tam_celdas =[1.5 * inch ]

    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')


##  Sedes  #
tituloModulo = "Reporte Módulo Sedes"
def getModuloSedes():
    query = """
            select 
            s.id as "ID",
            s.sede as "Sede", 
            s.direccion as "Dirección",
            s.descripcion as "Descripción",
            o.empresa as "Empresa"
            from sedes s 
            inner join organizacions o on s.organizacion_id=o.id 
            where s.deleted_at is null 
        """
    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@reports.get('/moduloSedes', tags=["ReportsXls"])
def getSedesExcel():
    resultados = getModuloSedes()
    fileRoute = DirectoryEmpleados + "sedes" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")
    
    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

@reports.get('/moduloSedes/pdf', tags=["ReportsPDF"])
def getSedesPDF():
    resultados = getModuloSedes()

    pdf_filename = f"moduloSedes_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["ID", "Sede", "Dirección", "Descripción", "Empresa"]
    tam_celdas =[1.5 * inch ]
    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')


## Niveles Jerarquicos  #
tituloModulo = "Reporte Módulo Niveles Jerarquicos"
def getNivelesJerarquicos():
    query = """
            select 
            pe.nombre as "Nivel", 
            descripcion as "Descripción" 
            from perfil_empleados pe  
            where pe.deleted_at is null
        """
    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@reports.get('/nivelesJerarquicos', tags=["ReportsXls"])
def getNivJerarquicosExcel():
    resultados = getNivelesJerarquicos()
    fileRoute = DirectoryEmpleados + "niveles-jerarquicos-" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")
    
    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)


@reports.get('/nivelesJerarquicos/pdf', tags=["ReportsPDF"])
def getNivelesJerarquicosPDF():
    resultados = getNivelesJerarquicos()

    pdf_filename = f"nivelesJerarquicos_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["Nivel", "Descripción"]
    tam_celdas =[1.5 * inch ]

    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')


## Registro de Áreas  #
tituloModulo = "Reporte Módulo Registro Áreas"
def getRegistroAreas():
    query = """
            select 
            a.id as "ID",
            a.area as "Nombre de área",
            g.nombre as "Grupo",
            r.area as "Reporta a",
            a.descripcion as "Descripción"
            from areas a 
            inner join grupos g on a.id_grupo=g.id
            left join areas r on a.id_reporta =r.id
            order by a.created_at asc
        """
    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@reports.get('/registroAreas', tags=["ReportsXls"])
def getRegAreasExcel():
    resultados=getRegistroAreas
    fileRoute = DirectoryEmpleados + "registroAreas-" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(
            status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

@reports.get('/registroAreas/pdf', tags=["ReportsPDF"])
def getRegAreasPDF():
    resultados = getRegistroAreas()
    pdf_filename = f"registroAreas_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["ID", "Nombre \nde área", "Grupo", "Reportar a", "Descripción"]
    tam_celdas =[1.5 * inch ]
    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')


## Macroprocesos  #
tituloModulo = "Reporte Módulo Macroprocesos"
def getMacroprocesos():
    query = """
            select 
            m.codigo as "Código",
            m.nombre as "Nombre",
            g.nombre as "Grupo" ,
            m.descripcion as "Descripción" 
            from macroprocesos m 
            inner join grupos g on m.id_grupo=g.id
            where m.deleted_at is null
            order by m.created_at asc
        """
    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@reports.get('/macroProcesos', tags=["ReportsXls"])
def getMacroProcesosExcel():
    resultados = getMacroprocesos()
    fileRoute = DirectoryEmpleados + "macroprocesos-" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

@reports.get('/macroProcesos/pdf', tags=["ReportsPDF"])
def getMacroProcesosPDF():
    resultados = getMacroprocesos()

    pdf_filename = f"macroProcesos_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["Código", "Nombre", "Grupo", "Descripción"]
    tam_celdas =[1.5 * inch ]

    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')



## Procesos  #
tituloModulo = "Reporte Módulo Procesos"
def getModuloProcesos():
    query = """
            select 
            p.codigo as "Código",
            p.nombre as "Nombre del proceso", 
            m.nombre as "Macroproceso",
            p.descripcion as "Descripción"
            from procesos p 
            inner join macroprocesos m on p.id_macroproceso=m.id 
            where p.deleted_at is null
            order by p.created_at asc
        """
    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@reports.get('/moduloProcesos', tags=["ReportsXls"])
def getModuloProcesosExcel():
    resultados = getModuloProcesos()
    fileRoute = DirectoryEmpleados + "moduloProcesos" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(
            status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

@reports.get('/moduloProcesos/pdf', tags=["ReportsPDF"])
def getModuloProcesosPDF():
    resultados = getModuloProcesos()
    pdf_filename = f"moduloProcesos_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["Código", "Nombre del \nproceso", "Macroproceso", "Descripción"]
    tam_celdas =[1.5 * inch ]

    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')


## Modulo Tipo Activos #
tituloModulo = "Reporte Módulo Tipo Activos"
def getModuloTipoActivos():
    query = """
            select 
            t.id as "ID",
            t.tipo as "Categoria"
            from tipoactivos t 
            where t.deleted_at is null
            order by t.created_at asc
        """
    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@reports.get('/moduloTipoActivos', tags=["ReportsXls"])
def getModuloTipoActivosExcel():
    resultados = getModuloTipoActivos()
    fileRoute = DirectoryEmpleados + "moduloTipoActivos" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(
            status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

@reports.get('/moduloTipoActivos/pdf', tags=["ReportsPDF"])
def getModuloTipoActivosPDF():
    resultados = getModuloTipoActivos()
    pdf_filename = f"moduloTipoActivos_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["ID", "Categoría"]
    tam_celdas =[1.5 * inch ]

    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')


## Modulo Activos  #
tituloModulo = "Reporte Módulo Activos"
def getModuloActivos():
    query = """
            select 
            t.id as "ID",
            t.tipo as "Categoria",
            sa.subcategoria as "Subcategoría"
            from tipoactivos t 
            inner join subcategoria_activos sa on t.id =sa.categoria_id  
            order by t.created_at asc
        """
    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@reports.get('/moduloActivos', tags=["ReportsXls"])
def getModuloActivosExcel():
    resultados = getModuloActivos()
    fileRoute = DirectoryEmpleados + "moduloActivos" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(
            status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

@reports.get('/moduloActivos/pdf', tags=["ReportsPDF"])
def getModuloActivosPDF():
    resultados = getModuloActivos()
    pdf_filename = f"moduloActivos_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["ID", "Caregoría", "Subcategoría"]
    tam_celdas =[1.5 * inch ]

    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')


## Inventario de Activos  #
tituloModulo = "Reporte Módulo Inventario Activos"
def getInventarioActivos():
    query = """
            select 
            a.id as "ID" ,
            a.nombreactivo as "Nombre del activo",
            t.tipo as "Categoria", 
            sa.subcategoria as "Subcategoría",
            a.descripcion as "Descripción",
            e.name as "Dueño",
            s.name as "Responsable"
            from tipoactivos t 
            inner join	subcategoria_activos sa on t.id=sa.categoria_id 
            inner join activos a on t.id =a.tipoactivo_id 
            inner join empleados e on a.dueno_id=e.id
            left join empleados s on e.supervisor_id=s.id 
        """
    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@reports.get('/inventarioActivos', tags=["ReportsXls"])
def getInventarioActivosExcel():
    resultados = getInventarioActivos()
    fileRoute = DirectoryEmpleados + "inventarioActivos" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(
            status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

@reports.get('/inventarioActivos/pdf', tags=["ReportsPDF"])
def getInventarioActivosPDF():
    resultados = getInventarioActivos()
    pdf_filename = f"inventarioActivos_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["ID", "Nombre \del activo", "Categoría", "Subcategoría", "Descripción", "Dueño", "Responsable"]
    tam_celdas =[1.5 * inch ]
    
    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')

    
## Glosario  #
tituloModulo = "Reporte Módulo Glosario"
def getGlosario():
    query = """
            select 
            g.numero as "Inciso",
            concepto as "Concepto", 
            norma as "Modulo",
            definicion as "Definición", 
            explicacion as "Explicación" 
            from glosarios g 
            where g.deleted_at is null 
        """
    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@reports.get('/glosario', tags=["ReportsXls"])
def getGlosarioExcel():
    resultados = getGlosario()
    fileRoute = DirectoryEmpleados + "glosario" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(
            status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

@reports.get('/glosario/pdf', tags=["ReportsPDF"])
def getGlosarioPDF():
    resultados = getGlosario()
    pdf_filename = f"glosario_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["Inciso", "Concepto", "Módulo", "Definición", "Explicación"]
    tam_celdas =[1.5 * inch ]

    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')


## Categorias capacitaciones  #
tituloModulo = "Reporte Módulo Categoria Capacitaciones"
def getCategoriasCapacitaciones():
    query = """
            select 
            distinct cc.id as "No.",
            cc.nombre as "Nombre"
            from recursos r 
            inner join categoria_capacitacions cc on r.categoria_capacitacion_id =cc.id
        """
    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@reports.get('/categoriasCapacitaciones', tags=["ReportsXls"])
def getCategoriasCapExcel():
    resultados = getCategoriasCapacitaciones()
    fileRoute = DirectoryEmpleados + "categoriasCapacitaciones" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(
            status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

@reports.get('/categoriasCapacitaciones/pdf', tags=["ReportsPDF"])
def getCategoriasCapacitacionesPDF():
    resultados = getCategoriasCapacitaciones()
    pdf_filename = f"categoriasCapacitaciones_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["No.", "Nombre"]
    tam_celdas =[1.5 * inch ]

    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')


## Logs  ## FALTA ##################
tituloModulo = "Reporte Módulo Visualizar Logs"
@reports.get('/visualizarLogs', tags=["ReportsXls"])
def getvisualizarLogs():
    query = """
        select 
        a.id as "ID",
        u.name as "User",
        a.event as "Event",
        a.old_values as "Old Value",
        a.new_values as "New Value",
        a.url as "Url",
        a.tags as "Tags",
        a.created_at as "Fecha creación",
        a.updated_at as "Fecha última actualización"
        from audits a 
        inner join users u on a.user_id =u.id 
        where u.deleted_at is null
        order by a.created_at desc
        """
    resultados = ejecutar_consulta_sql(cursor, query)
    
    if not resultados:
        raise HTTPException(status_code=404, detail="No data found for the query")
    

    df = pd.DataFrame(resultados)
    
    def limpiar_json(columna):
        def extraer_datos(json_str):
            try:
                data = json.loads(json_str)
                result = f"id: {data['id']}, especificaciones: {data['especificaciones']}, cantidad: {data['cantidad']}"
                return result
            except (json.JSONDecodeError, KeyError, TypeError) as e:
                print(f"Error procesando JSON: {e}")
                return json_str
        return columna.apply(extraer_datos)
    
    # Verificar si las columnas 'Old Value' y 'New Value' existen en el DataFrame antes de limpiarlas
    if 'Old Value' in df.columns:
        df['Old Value'] = limpiar_json(df['Old Value'])
    else:
        print("'Old Value' column not found in the data")
        raise HTTPException(status_code=500, detail="'Old Value' column not found in the data")
    
    if 'New Value' in df.columns:
        df['New Value'] = limpiar_json(df['New Value'])
    else:
        print("'New Value' column not found in the data")
        raise HTTPException(status_code=500, detail="'New Value' column not found in the data")
    
    fileRoute = DirectoryEmpleados + "visualizarLogs" + str(now) + ".xlsx"
    df.to_excel(fileRoute, index=False)
    exportar_a_excel(resultados, fileRoute)
    ajustar_columnas(fileRoute)
    excel_path = Path(fileRoute)
    
    if not excel_path.is_file():
        raise HTTPException(status_code=404, detail="file not found on the server")
    
    return FileResponse(excel_path)
#################


## Registro Timesheet           #
tituloModulo = "Reporte Módulo Registro Timesheet"
def getRegistroTimesheet(
    area: Optional[str] = None,
    empleado: Optional[str] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None
):
    if fecha_inicio and fecha_fin:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(
                status_code=400, detail="Formato de fecha incorrecto. Use 'YYYY-MM-DD'.")

    query = """
        select
        to_char(date_trunc('week', t.fecha_dia), 'DD/MM/YYYY') as "Fecha inicio",
        to_char(t.fecha_dia, 'DD/MM/YYYY') as "Fecha fin",
        e.name as "Empleado",
        p.name as "Aprovador",
        a.area as "Área",
        t.estatus as "Estatus",
        sum(
        coalesce(cast(nullif(th.horas_lunes, '') as numeric), 0) +
        coalesce(cast(nullif(th.horas_martes, '') as numeric), 0) +
        coalesce(cast(nullif(th.horas_miercoles, '') as numeric), 0) +
        coalesce(cast(nullif(th.horas_jueves, '') as numeric), 0) +
        coalesce(cast(nullif(th.horas_viernes, '') as numeric), 0) +
        coalesce(cast(nullif(th.horas_sabado, '') as numeric), 0) +
        coalesce(cast(nullif(th.horas_domingo, '') as numeric), 0)
        ) as "Horas de la semana"
        from timesheet t
        inner join empleados e ON t.empleado_id = e.id
        inner join empleados p ON t.aprobador_id = p.id
        inner join areas a on e.id = a.empleados_id
        inner join timesheet_horas th on t.id = th.timesheet_id
        where e.deleted_at is null
    """

    if area:
        query += f" and a.area = '{area}'"
    if empleado:
        query += f" and e.name = '{empleado}'"
    if fecha_inicio and fecha_fin:
        query += f" and tgit.fecha_dia between '{fecha_inicio}' and '{fecha_fin}'"

    query += """
        group by
            t.fecha_dia,
            e.name,
            p.name,
            a.area,
            t.estatus
        order by t.fecha_dia desc;
    """

    #print(query)

    file_path = "query.txt"
    with open(file_path, "w") as file:
        file.write(query)

    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@reports.post('/registrosTimesheet/', tags=["ReportsXls"])
def getRegistrosTimesheetExcel():
    resultados=getRegistroTimesheet()
    fileRoute = DirectoryEmpleados + "registroTimesheet" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(
            status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

@reports.post('/registrosTimesheet/pdf', tags=["ReportsPDF"])
def getRegistrosTimesheetPDF():
    resultados=getRegistroTimesheet()
    pdf_filename = f"registroTimesheet_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["Fecha inicio", "Fecha fin", "Empleado", "Aprovador", "Área", "Estatus", "Horas de \nla semana"]
    tam_celdas =[1.5 * inch ]

    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')


## Timesheet Áreas  #
tituloModulo = "Reporte Módulo Timesheet Áreas"
def getTimesheetAreas(
    area: Optional[str] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None
):
    if fecha_inicio and fecha_fin:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(
                status_code=400, detail="Formato de fecha incorrecto. Use 'YYYY-MM-DD'.")

    query = """
            select 
            e.name as "Nombre",
            p.puesto as "Puesto",
            a.area as "Área",
            e.estatus as "Estatus",
            e.antiguedad as "Fecha"
            from empleados e 
            inner join puestos p on e.puesto_id =p.id 
            inner join areas a on e.area_id=a.id
            inner join timesheet t on e.id=t.empleado_id 
        """
    if area:
        query += f" and a.area = '{area}'"
    if fecha_inicio and fecha_fin:
        query += f" and t.fecha_dia between '{fecha_inicio}' and '{fecha_fin}'"

    query += """
        group by
            a.area,
            e.name,
            p.puesto,
            e.estatus,
            e.antiguedad,
            t.fecha_dia
    """   

    file_path = "query.txt"
    with open(file_path, "w") as file:
        file.write(query)

    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@reports.post("/timesheetAreas/", tags=["ReportsXls"])
def getTimesheetAreasExcel():
    resultados = getTimesheetAreas()
    fileRoute = DirectoryEmpleados + "timesheetAreas" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(
            status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

@reports.post("/timesheetAreas/pdf", tags=["ReportsPDF"])
def getTimesheetAreasPDF():
    resultados = getTimesheetAreas()
    pdf_filename = f"timesheetAreas_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["Nombre", "Puesto","Área","Estatus", "Fecha"]
    tam_celdas =[1.5 * inch ]

    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')



## Timesheet proyectos  #
tituloModulo = "Reporte Módulo Timesheet Proyectos"
def getTimesheetProyectos(
    area: Optional[str] = None,
    proyecto: Optional[str] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None
    ):

    if fecha_inicio and fecha_fin:   
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(
                status_code=400, detail="Formato de fecha incorrecto. Use 'YYYY-MM-DD'.")

    query = """
            select 
            tp.proyecto as "ID-Proyecto",
            string_agg(distinct a.area, ', ') as "Áreas participantes",
            string_agg(distinct e.name, ', ') as "Empleados participantes",
            tc.nombre as "Cliente"
            from timesheet_proyectos tp 
            left join timesheet_proyectos_empleados tpe on tp.id=tpe.proyecto_id 
            left join timesheet_proyectos_areas tpa on tp.id =tpe.proyecto_id 
            left join areas a on tpe.area_id =a.id  
            left join empleados e on tpe.empleado_id =e.id 
            right  join timesheet_clientes tc on tp.cliente_id =tc.id           
        """
    
    if area:
        query += f" and a.area = '{area}'"
    if proyecto:
        query += f" and tp.proyecto = '{proyecto}'"
    if fecha_inicio and fecha_fin:
        query += f" and tp.fecha_inicio between '{fecha_inicio}' and '{fecha_fin}'"


    query += """
        group by 
            a.area,
            tp.proyecto, 
            tc.nombre,
            tp.fecha_inicio 
    """   

    file_path = "query.txt"
    with open(file_path, "w") as file:
        file.write(query)

    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@reports.post('/timesheetProyectos/', tags=["ReportsXls"])
def getTimesheetProyectosExcel():
    resultados = getTimesheetProyectos()
    fileRoute = DirectoryEmpleados + "timesheetProyectos" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(
            status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

@reports.post('/timesheetProyectos/pdf', tags=["ReportsPDF"])
def getTimesheetProyectosPDF():
    resultados = getTimesheetProyectos()
    pdf_filename = f"timesheetProyectos_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["ID-Proyecto", "Áreas \nParticipantes", "Empleados  \nParticipantes", "Cliente"]
    tam_celdas =[1.5 * inch ]

    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')


## Registros Colaboradores Tareas   #
tituloModulo = "Reporte Módulo Colaboradores Tareas"
def getColaboradoresTareas(
    empleado: Optional[str] = None,
    proyecto: Optional[str] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None
    ):

    if fecha_inicio and fecha_fin:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(
                status_code=400, detail="Formato de fecha incorrecto. Use 'YYYY-MM-DD'.")

    query = """
            select 
            tp.fecha_inicio as "Fecha inicio", 
            tp.fecha_fin as "Fecha fin",
            string_agg(distinct e.name, ', ') as  "Empleado",
            string_agg(distinct s.name, ', ') as "Supervisor",
            string_agg(distinct tp.proyecto, ', ') as "Proyecto",
            string_agg(distinct tt.tarea, ', ') as "Tarea",
            th.descripcion as "Descripción",
            sum(
                coalesce(cast(nullif(th.horas_lunes, '') as numeric), 0) +
                coalesce(cast(nullif(th.horas_martes, '') as numeric), 0) +
                coalesce(cast(nullif(th.horas_miercoles, '') as numeric), 0) +
                coalesce(cast(nullif(th.horas_jueves, '') as numeric), 0) +
                coalesce(cast(nullif(th.horas_viernes, '') as numeric), 0) +
                coalesce(cast(nullif(th.horas_sabado, '') as numeric), 0) +
                coalesce(cast(nullif(th.horas_domingo, '') as numeric), 0)
            ) as "Horas de la semana"
            from timesheet_proyectos tp 
            left join timesheet_proyectos_empleados tpe on tp.id =tpe.proyecto_id 
            left join empleados e on tpe.empleado_id =e.id 
            left join empleados s on e.supervisor_id=s.id
            left join timesheet_tareas tt on tp.id =tt.proyecto_id
            right join timesheet_horas th on e.id=th.empleado_id 
            where tp.fecha_inicio > '2022-01-01'
        """
    if empleado:
        query += f" and e.name = '{empleado}'"
    if proyecto:
        query += f" and tp.proyecto = '{proyecto}'"
    if fecha_inicio and fecha_fin:
        query += f" and t.fecha_inicio between '{fecha_inicio}' and '{fecha_fin}'"

    query += """
        group by
            e.name,
            tp.proyecto,
            tp.fecha_inicio, 
            tp.fecha_fin,
            th.descripcion 
        order by tp.fecha_inicio desc;
    """

    print(query)

    file_path = "query.txt"
    with open(file_path, "w") as file:
        file.write(query)

    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados


@reports.post('/colaboradoresTareas/', tags=["ReportsXls"])
def getColaboradoresTareasExcel():
    resultados = getColaboradoresTareas()
    fileRoute = DirectoryEmpleados + "colaboradoresTareas" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(
            status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

@reports.post('/colaboradoresTareas/pdf', tags=["ReportsPDF"])
def getColaboradoresTareasPDF():
    resultados = getColaboradoresTareas()
    pdf_filename = f"colaboradoresTareas_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["Fecha inicio", "Fecha fin", "Empleado", "Supervisor", "Proyecto", "Tarea", "Descripción", "Horas de \nla semana"]
    tam_celdas =[1.5 * inch ]
    
    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')


## Timesheet Financiero        #
tituloModulo = "Reporte Módulo Timesheet Financiero"
def getTimesheetFinanciero(
    proyecto: Optional[str] = None):

    query = """
            select
            tp.identificador as "ID",
            tp.proyecto as "Proyecto",
            tc.nombre as "Cliente",
            a.area as "Área(s)",
            e.name as "Empleados participantes",
            tpe.horas_asignadas as "Horas del empleado",
            tpe.horas_asignadas * tpe.costo_hora as "Costo total del empleado",
            tp.estatus as "Estatus",
                sum(tpe.horas_asignadas)over(partition by tpe.proyecto_id) as "Horas totales del proyecto",
                sum(tpe.horas_asignadas * tpe.costo_hora) over(partition by tpe.proyecto_id) as "Costo total del Proyecto"
            from timesheet_proyectos tp 
            left join timesheet_clientes tc on tp.cliente_id =tc.id
            left join timesheet_proyectos_empleados tpe on tp.id =tpe.proyecto_id 
            left join areas a on tpe.area_id =a.id 
            left join empleados e on tpe.empleado_id =e.id 
        """
    
    if proyecto:
        query += f" and tp.proyecto = '{proyecto}'"

    query += """
        group by
            tp.proyecto,
            tp.identificador,
            tc.nombre,
            a.area,
            e.name,
            tpe.horas_asignadas,
            tpe.costo_hora,
            tp.estatus,
            tpe.proyecto_id
    """    

    file_path = "query.txt"
    with open(file_path, "w") as file:
        file.write(query)

    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@reports.post('/timesheetFinanciero/', tags=["ReportsXls"])
def getTimesheetFinancieroExcel():
    resultados = getTimesheetFinanciero()
    fileRoute = DirectoryEmpleados + "timesheetFinanciero" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")
    
    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(
            status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)


@reports.post('/timesheetFinanciero/pdf', tags=["ReportsPDF"])
def getTimesheetFinancieroPDF():
    resultados = getTimesheetFinanciero()
    pdf_filename = f"timesheetFinanciero_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["ID", "Proyecto", "Cliente", "Área(s)", "Empleados \nParticipantes", "Horas del \nempleado", "Costo total \ndel empleado", "Empleado", "Estatus", "Horas totales \ndel proyecto", "Costo total \ndel proyecto"]
    tam_celdas =[1.5 * inch ]

    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')


## Vista global de Solicitudes de Day Off 
tituloModulo = "Reporte Módulo Vista Global  Solicitudes de Day Off" 
def getSolicitudesDayOff():
    query = """
            select e.name as "Solicitante",
            sd.descripcion as "Descripcion",
            to_char(sd.año, 'FM9999') as "Año",
            sd.dias_solicitados as "Días solicitados",
            sd.fecha_inicio as "Inicio",
            sd.fecha_fin as "Fin",
            case  
                when sd.aprobacion = 3 then 'Aprobado'
                when sd.aprobacion = 2 then 'Rechazado'
                when sd.aprobacion = 1 then 'Pendiente'
                else 'desconocido'
            end as "Aprobacion" 
            from solicitud_dayoff sd 
            inner join empleados e on sd.empleado_id =e.id
            order by sd.año desc 
        """
    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@reports.get('/solicitudesDayOff', tags=["ReportsXls"])
def getSolicitudesDayOffExcel():
    resultados = getSolicitudesDayOff()
    fileRoute = DirectoryEmpleados + "solicitudesDayOff" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

@reports.get('/solicitudesDayOff/pdf', tags=["ReportsPDF"])
def getSolicitudesDayOffPDF():
    resultados = getSolicitudesDayOff()
    pdf_filename = f"solicitudesDayOff_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["Solicitante", "Descripción", "Año", "Diás \nsolicitados", "Inicio", "Fin", "Aprobación"]
    tam_celdas =[1.5 * inch ]

    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')


## Vista GLobal Solicitudes de Vacaciones   #
tituloModulo = "Reporte Módulo Vista Global Solicitudes Vacaciones"
def getSolicitudesVacaciones():
    query = """
            select e.name as "Solicitante",
            sv.descripcion as "Descripción",
            sv.año as "Periodo",
            sv.dias_solicitados  as "Días solicitados",
            sv.fecha_inicio as "Inicio",
            sv.fecha_fin as "Fin",
            case  
                when sv.aprobacion = 3 then 'Aprobado'
                when sv.aprobacion = 2 then 'Rechazado'
                when sv.aprobacion = 1 then 'Pendiente'
                else 'desconocido'
            end as "Aprobacion"
            from solicitud_vacaciones sv 
            inner join empleados e on sv.empleado_id =e.id 
            order by sv.fecha_inicio desc 
        """
    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@reports.get('/solicitudesVacaciones', tags=["ReportsXls"])
def getSolicitudesVacacionesExcel():
    resultados = getSolicitudesVacaciones()
    fileRoute = DirectoryEmpleados + "solicitudesVacaciones" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(
            status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

@reports.get('/solicitudesVacaciones/pdf', tags=["ReportsPDF"])
def getSolicitudesVacacionesPDF():
    resultados = getSolicitudesVacaciones()
    pdf_filename = f"solicitudesVacaciones_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["Solicitante", "Descripción", "Periodo", "Días  \nsolicitados", "Inicio", "Fin", "Aprobación"]
    tam_celdas =[1.5 * inch ]

    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')


## Evaluaciones 360     #
tituloModulo = "Reporte Módulo Evaluaciones 360"
def getEvaluaciones360():
    query = """
            select id as "ID",
            nombre as "Nombre",
            case  
                when estatus:: integer = 3 then 'Cerrado'
                when estatus:: integer = 2 then 'Abierto'
                when estatus:: integer = 1 then 'Pendiente'
                else 'desconocido'
            end as "Estatus",
            fecha_inicio as "Fecha inicio",
            fecha_fin as "Fecha fin",
            case 
                when include_competencias then 'si'
                else 'no'
            end  as "¿Incluye competencias?",
            case 
                when include_objetivos then 'si'
                else 'no'
            end  as "¿Incluye objetivos?"
            from ev360_evaluaciones ee 
            where estatus ::integer = 3
                and include_competencias = true 
                and include_objetivos = true 
        """
    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados
    
@reports.get('/evaluaciones360', tags=["ReportsXls"])
def getEvaluaciones360Excel():
    resultados = getEvaluaciones360()
    fileRoute = DirectoryEmpleados + "evaluaciones360" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(
            status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

@reports.get('/evaluaciones360/pdf', tags=["ReportsPDF"])
def getEvaluaciones360PDF():
    resultados = getEvaluaciones360()
    pdf_filename = f"evaluaciones360_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["ID", "Nombre", "Estatus", "RFecha inicio", "Fecha fin", "¿Incluye competencias?", "¿Incluye objetivos?"]
    tam_celdas =[1.5 * inch ]

    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')


## Empleados controller     #
tituloModulo = "Reporte Módulo Empleados Controller"
def getEmpleadoController(
    empleado: Optional[str] = None
    ):
    
    query = """
            select 
            e.name as "Empleado",
            p.name as "Supervisor",
            s.sede as "Sede",
            pe.nombre as "Perfil",
            string_agg(distinct ce.nombre, ', ' ) as "Certificaciones",
            ee.institucion as "Educación"
            from empleados e 
            left join empleados p on e.supervisor_id=p.id 
            inner join sedes s on e.sede_id=s.id
            left join perfil_empleados pe on e.perfil_empleado_id=pe.id 
            left join certificaciones_empleados ce on e.id=ce.empleado_id 
            left join educacion_empleados ee on e.id=ee.empleado_id 
            where e.estatus= 'alta'
        """
        
    if empleado:
            query += f" and e.name = '{empleado}'"

    query += """
        group by
            e.name,
            p.name,
            s.sede,
            pe.nombre, 
            ee.institucion 
        order by e.name asc;
    """
    file_path = "query.txt"
    with open(file_path, "w") as file:
        file.write(query)

    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@reports.post('/empleadosController/', tags=["ReportsXls"])
def getEmpleadoControllerExcel():
    resultados = getEmpleadoController()
    fileRoute = DirectoryEmpleados + "empleadoController" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(
            status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

@reports.post('/empleadosController/pdf', tags=["ReportsPDF"])
def getEmpleadoControllerPDF():
    resultados = getEmpleadoController()
    pdf_filename = f"empleadosController_{now}.pdf"
    pdf_path = Path(DirectoryEmpleados) / pdf_filename

    encabezados = ["Empleado", "Supervisor", "Sede", "Perfil", "Certificaciones", "Educación"]
    tam_celdas =[1.5 * inch ]

    try:
        generar_pdf_generalizado(resultados, str(pdf_path), encabezados, tam_celdas, orientacion='horizontal')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")

    # Verificar si el archivo se creó correctamente
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF file not found on the server")

    # Devolver el archivo para descarga
    return FileResponse(path=str(pdf_path), filename=pdf_filename, media_type='application/pdf')


########


def ejecutar_consulta_sql(cursor, consulta):
    try:
        cursor.execute(consulta)
        resultados = cursor.fetchall()
        return resultados
    except psycopg2.Error as e:
        print("Error al ejecutar la consulta SQL:" + str(e))
        # return JSONResponse(content={"message": "Error al ejecutar la consulta SQL."})
        raise HTTPException(
            status_code=500, detail="Error executing SQL query: " + str(e))


def exportar_a_excel(resultados, nombre_archivo):
    try:
        if resultados is not None:
            df = pd.DataFrame(resultados, columns=[
                desc[0] for desc in cursor.description])
            df.to_excel(nombre_archivo, index=False)
            print("Resultados exportados a", nombre_archivo)
    except Exception as e:
        print("No se pudieron exportar los resultados a Excel debido a un error." + str(e))
        raise HTTPException(
            status_code=500, detail="Report error: " + str(e))

def ajustar_columnas(nombre_archivo):
    try:
        workbook = load_workbook(nombre_archivo)
        worksheet = workbook.active

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

        workbook.save(nombre_archivo)
        print("Columnas ajustadas en", nombre_archivo)
    except Exception as e:
        print("No se pudieron ajustar las columnas debido a un error." + str(e))
        raise HTTPException(
            status_code=500, detail="Column adjust error: " + str(e))
    
def generar_pdf_generalizado(datos, nombre_archivo, encabezados, tam_celdas , orientacion='horizontal'):
    # Definir la orientación de la página
    if orientacion == 'horizontal':
        pagesize = landscape(letter)
    else:
        pagesize = letter

    # Crear el documento PDF
    doc = SimpleDocTemplate(nombre_archivo, pagesize=pagesize)
    elementos = []

    estilo_titulo = ParagraphStyle(
        name='TituloTabla',
        fontName='Roboto',
        fontSize=16,
        textColor=HexColor("#2E2E2E"),
        alignment=1,  
        spaceAfter=12,
        borderPadding= 0 
    )

    estilo_encabezado = ParagraphStyle(
        name='Encabezado',
        fontName='Roboto',
        fontSize=11,
        textColor=HexColor("#3D3D3D"),
        alignment=0  
    )

    estilo_encabezadoDerecha = ParagraphStyle(
        name='Encabezado2',
        fontName='Roboto',
        fontSize=13,
        textColor=HexColor("#49598A"),
        alignment=2,
        leftIndent= 2  
    )

    estilo_normal = ParagraphStyle(
        name='Roboto',
        fontName='Roboto',
        fontSize=11,
        textColor= HexColor("#2E2E2E"),
        alignment=0,
        spaceBefore=3  
    )

    logo_path = "silent.png"  # Asegúrate de proporcionar la ruta correcta
    logo = Image(logo_path, width=1.5 * inch, height=1 * inch, hAlign='LEFT')  # Ajusta el tamaño del logo

    info_empresa = Paragraph("SILENT4BUSINESS. S.A. DE C.V.<br/> RFC:SIL160727HV7<br/>INSURGENTES SUR 2453 TIZAPÁN SAN ÁNGEL <br/> ÁLVARO OBREGÓN, CIUDAD DE MÉXICO. C.P.01090", estilo_encabezado)
    
    info_derecha = Paragraph("Reporte <br/> Fecha:", estilo_encabezadoDerecha)
    info_derecha_bg = Table([[info_derecha]], colWidths=[2.5 * inch], rowHeights=[1 * inch])
    info_derecha_bg.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), HexColor("#EEFCFF")),  # Fondo azul claro
        ('VALIGN', (0, 0), (-1, -1), 'TOP')
    ]))

    encabezado_tabla = Table(
        [[logo, info_empresa, info_derecha_bg]],
        colWidths=[1.8 * inch, 4 * inch, 2.5 * inch]
    )
    encabezado_tabla.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0)
    ]))

    elementos.append(encabezado_tabla)
    elementos.append(Spacer(0, 0.5 * inch))  # Espacio entre el encabezado y el título

    titulo = Paragraph(tituloModulo, estilo_titulo)
    elementos.append(titulo)

    datos_tabla = [encabezados] + [
        [Paragraph(str(item), estilo_normal) for item in fila]
        for fila in datos
    ]
    tabla = Table(datos_tabla, colWidths=tam_celdas, repeatRows=1)
    # tabla = Table(datos_tabla, repeatRows=1)

    color_encabezado = HexColor("#D8F2FF") #Azul medio
    color_fila_par = HexColor("#E9E9E9")  # Gris claro
    color_fila_impar =  HexColor("#FFFFFF")    #Blanco
    color_texto_encabezado = HexColor("#575757")  #Gris medio
    color_borde = HexColor('#CCCCCC')

    # Estilo de la tabla
    estilo_tabla = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), color_encabezado),
        ('TEXTCOLOR', (0, 0), (-1, 0), color_texto_encabezado),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Roboto'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 11),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Roboto'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('TOPPADDING', (0, 1), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 11),
        ('GRID', (0, 0), (-1, -1), 1, color_borde),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBEFORE', (0, 0), (-1,0), 0, '#D8F2FF'),
        ('LINEAFTER', (0, 0), (-1,0), 0, '#D8F2FF')
    ])

    for i, fila in enumerate(datos_tabla[1:], start=1):
        bg_color = color_fila_par if i % 2 == 0 else color_fila_impar
        estilo_tabla.add('BACKGROUND', (0, i), (-1, i), bg_color)

    tabla.setStyle(estilo_tabla)

    # num_columnas = len(encabezados)
    # for i in range(num_columnas):
    #     tabla._argW[i] = 1.5 * inch  


    elementos.append(tabla)

    # Generar el PDF
    doc.build(elementos)

