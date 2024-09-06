import pandas as pd
from config.database import cursor
from fastapi import FastAPI, Query, HTTPException
from fastapi.responses import FileResponse
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import psycopg2
from fastapi import APIRouter
from pathlib import Path
import os
from datetime import date, datetime

app = FastAPI()
cert = APIRouter()
DirectoryEmpleados = "reportsfile/administracion/empleados/"
# Validar si la carpeta ya existe
if not os.path.exists(DirectoryEmpleados):
    # Si no existe, crear la carpeta
    os.makedirs(DirectoryEmpleados)

now = date.today()

def getCertificados():
    query = """
            select 
            e.name as "Nombre Completo",
            a.area as "Área",
            ce.nombre as "Certificados",
            ce.vigencia as "Vigencia"
            from empleados e 
            inner join areas a  on e.id =a.empleados_id 
            left join certificaciones_empleados ce on e.id=ce.empleado_id 
            order by e.name asc 
        """
    resultados = ejecutar_consulta_sql(cursor, query)
    return resultados

@cert.post('/certificados/', tags=["ReportsXlss"])
def getCertificadosExcel():
    resultados = getCertificados()
    fileRoute = DirectoryEmpleados + "certificados" + str(now) + ".xlsx"
    try:
        exportar_a_excel(resultados, fileRoute)
        ajustar_columnas(fileRoute)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

    excel_path = Path(fileRoute)
    if not excel_path.is_file():
        raise HTTPException(
            status_code=404, detail="file not found on the server")
    return FileResponse(excel_path)

def ejecutar_consulta_sql(cursor, consulta):
    try:
        cursor.execute(consulta)
        resultados = cursor.fetchall()
        return resultados
    except psycopg2.Error as e:
        print("Error al ejecutar la consulta SQL:" + str(e))
        # return JSONResponse(content={"message": "Error al ejecutar la consulta SQL."})
        raise HTTPException(
            status_code=500, detail="Error executing SQL query: " + str(e))


def limpiar_datos(df):
    try:
        # 1. Eliminar filas duplicadas
        df.drop_duplicates(inplace=True)

        # 2. Eliminar filas con valores nulos en todas las columnas
        df.dropna(how='all', inplace=True)

        # 3. Rellenar valores nulos con un valor predeterminado en columnas específicas
        df.fillna({"Área": "No especificada", "Certificados": "Sin Certificado"}, inplace=True)

        # 4. Eliminar espacios en blanco en los extremos de los textos
        df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)

        return df
    except Exception as e:
        print("Error durante la limpieza de datos: " + str(e))
        raise HTTPException(status_code=500, detail="Data cleaning error: " + str(e))

def exportar_a_excel(resultados, nombre_archivo):
    try:
        if resultados is not None:
            # Convertir los resultados a un DataFrame de pandas
            df = pd.DataFrame(resultados, columns=[desc[0] for desc in cursor.description])
            
            # Limpiar los datos antes de exportar
            df_limpio = limpiar_datos(df)
            
            # Exportar el DataFrame limpio a un archivo Excel
            df_limpio.to_excel(nombre_archivo, index=False)
            print("Resultados exportados a", nombre_archivo)
    except Exception as e:
        print("No se pudieron exportar los resultados a Excel debido a un error." + str(e))
        raise HTTPException(status_code=500, detail="Report error: " + str(e))
    
def ajustar_columnas(nombre_archivo):
    try:
        workbook = load_workbook(nombre_archivo)
        worksheet = workbook.active

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

        workbook.save(nombre_archivo)
        print("Columnas ajustadas en", nombre_archivo)
    except Exception as e:
        print("No se pudieron ajustar las columnas debido a un error." + str(e))
        raise HTTPException(
            status_code=500, detail="Column adjust error: " + str(e))
